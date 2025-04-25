import streamlit as st
import openpyxl
from openpyxl.drawing.image import Image
import os

IMAGE_FOLDER = 'uploaded_images'
TEMP_EXCEL_FILE = 'temp_data_gambar.xlsx'

# Buat folder gambar jika belum ada
if not os.path.exists(IMAGE_FOLDER):
    os.makedirs(IMAGE_FOLDER)

st.title("📸 Upload Gambar + Keterangan")

# 🔴 Tambahkan Tombol Reset Data
if st.button("🔄 Reset Data"):
    if os.path.exists(TEMP_EXCEL_FILE):
        os.remove(TEMP_EXCEL_FILE)
    if os.path.exists(IMAGE_FOLDER):
        for file in os.listdir(IMAGE_FOLDER):
            os.remove(os.path.join(IMAGE_FOLDER, file))
    st.success("Data lama berhasil dihapus! Mulai dari awal.")

# Upload file Excel lama (opsional)
uploaded_excel = st.file_uploader("Upload file Excel sebelumnya (optional)", type=["xlsx"])

if uploaded_excel:
    with open(TEMP_EXCEL_FILE, "wb") as f:
        f.write(uploaded_excel.getbuffer())
    st.success("File Excel lama berhasil dimuat!")
else:
    if not os.path.exists(TEMP_EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        wb.save(TEMP_EXCEL_FILE)

# Upload 1 gambar + keterangan
uploaded_file = st.file_uploader("Upload 1 gambar", type=["jpg", "png", "jpeg"])
description = st.text_input("Masukkan keterangan gambar")

if st.button("💾 Simpan ke Excel"):
    if uploaded_file and description:
        wb = openpyxl.load_workbook(TEMP_EXCEL_FILE)
        ws = wb.active

        ws['A1'] = "Gambar"
        ws['B1'] = "Keterangan"

        # Hitung jumlah gambar yang sudah ada
        existing_images = len(ws._images)
        next_row = (existing_images * 1) + 2     

        # Simpan gambar
        image_path = os.path.join(IMAGE_FOLDER, uploaded_file.name)
        with open(image_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        img = Image(image_path)

        # ✅ Tetapkan ukuran gambar sedang (180x180 px)
        img.width = 180
        img.height = 180

        # Sisipkan gambar di cell sesuai kelipatan 5
        ws.add_image(img, f'A{next_row}')

        # Atur ukuran cell agar gambar rapi
        ws.row_dimensions[next_row].height = 140
        ws.column_dimensions['A'].width = 25

        # Tambahkan keterangan di kolom B
        ws[f'B{next_row}'] = description

        wb.save(TEMP_EXCEL_FILE)
        st.success(f"Gambar disimpan di A{next_row}, keterangan di B{next_row}")
    else:
        st.error("Silakan upload gambar dan isi keterangan terlebih dahulu.")

# Download file Excel terbaru
if os.path.exists(TEMP_EXCEL_FILE):
    with open(TEMP_EXCEL_FILE, "rb") as f:
        st.download_button('⬇️ Download Excel', f, file_name='data_gambar_final.xlsx')
